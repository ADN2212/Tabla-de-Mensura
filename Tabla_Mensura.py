import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from io import open
from openpyxl import Workbook
from math import atan
from math import pi
from math import hypot
from openpyxl.styles import colors, Font, Border, Side, Alignment
import re

ventana = tk.Tk()
ventana.geometry("260x115")
ventana.title("Tabla Mensura")
ventana.resizable(False, False)
#-----------------------------------------------------------------------------Variables------------------------------------------------------------------------------------------------

formato_elegido = tk.IntVar()#su valor inicial es cero por definicion.
puntos = []

#--------------------------------------------------------------------------Funciones y Clase-------------------------------------------------------------------------------------------

class Punto:	
	def __init__(self, nombre, este, norte):
		self.nombre = nombre
		self.este = este
		self.norte = norte

def calcular_GMS(angulo, c):
	"""
	Esta función recibe el angulo en grados sexagesimales decimales.
	El parametro c indaca hasta donde llegara el angulo, 1 = solo grados, 2 = grados y minutos, 3 = grados, minutos y segundos.
	"""
	if angulo < 0:
		angulo = -1 * angulo

	entero = 0
	decimal = 0
	GMS = [0, 0, 0]

	for i in range(3):
		entero = int(angulo)
		decimal = angulo - entero

		if entero < 10:
			GMS[i] = "0" + str(entero)#Para agragar el cero delante, esto es con fines esteticos.
		else:
			GMS[i] = entero

		angulo = decimal * 60

	if c == 3:	
		return str(GMS[0]) + "°" + str(GMS[1]) + "'" + str(GMS[2]) + '"'

	elif c == 2:
		return str(GMS[0]) + "°" + str(GMS[1]) + "'" 
	
	elif c == 1:
		return str(GMS[0]) + "°"

	else:
		return None		



def rumbo_y_dist(punto_1, punto_2):
	"""Esta funcion recibe dos objetos de tipo Punto para retornar una tupla con la distancia y el rumbo"""

	delta_x = punto_2.este - punto_1.este
	delta_y = punto_2.norte - punto_1.norte
	distancia = round(hypot(delta_x, delta_y), 2)
	rumbo = "No definido"
#Es bueno enlazarlos para que una vez se cumpla una ignore las demas

	if delta_x > 0 and delta_y > 0:
		rumbo = "N {} E".format(calcular_GMS(atan(delta_x/delta_y)*(180/pi), 2))

	elif delta_x > 0 and delta_y < 0:
		rumbo = "S {} E".format(calcular_GMS(atan(delta_x/delta_y)*(180/pi), 2))

	elif delta_y < 0 and delta_y < 0:
		rumbo = "S {} W".format(calcular_GMS(atan(delta_x/delta_y)*(180/pi), 2))	

	elif delta_x < 0 and delta_y > 0:
		rumbo = "N {} W".format(calcular_GMS(atan(delta_x/delta_y)*(180/pi), 2))

	#Los siguientes casos son menos comunes:

	elif delta_x > 0 and delta_y == 0:
		rumbo = "N 90°00' E"#tambien puede ser "S 90°00' E"
	
	elif delta_x == 0 and delta_y < 0:
		rumbo = "S 00°00' E"#tambien puede ser "S 00° 00' W"

	elif delta_x < 0 and delta_y == 0:
		rumbo = "N 90°00' W"#tambien puede ser "S 90° 00' W"
	
	elif delta_x == 0 and delta_y > 0:
		rumbo = "N 00°00' E"#tambien puede ser "N 00° 00' W"

	return (rumbo, round(distancia, 2))

def mostrar_intrucciones():
	global formato_elegido

	if formato_elegido.get() == 1:
		messagebox.showinfo("PM", "Inserte el archivo de tipo txt que genera AutoCAD cuando se aplica el comando 'PM' a una polilinea.")

	elif formato_elegido.get() == 2:
		text1 = "Inserte un archivo de texto (.txt) que contenga la información \n"
		text2 = "que retorna la linea de comados de AutoCAD cuando se aplica el coamdo 'list'"
		text3 = "a una polilinea.\n -Ojo: -->SOLO LAS COORDENADAS<--"

		messagebox.showinfo("List", text1 + text2 + text3)
	
	else:
		messagebox.showwarning("???", "Por favor elija el formato del cual quiere recibir las Instrucciones")			

def extraerXY(linea):
	if "X" in linea and "Y" in linea and "Z" in linea and "at point" in linea:
		pos_x = re.search("X", linea).span()[0]
		pos_y = re.search("Y", linea).span()[0]
		pos_z = re.search("Z", linea).span()[0]
		
		Este = ""
		Norte = ""
		
		for i in range(pos_x, pos_y):
			if linea[i].isdigit() or linea[i] == ".":
				Este += linea[i]

		for i in range(pos_y, pos_z):
			if linea[i].isdigit() or linea[i] == ".":
				Norte += linea[i]	

		return (float(Este), float(Norte))
				
	else:	
		return False

def cargar():
	global puntos
	global formato_elegido
	pasaron = True#sera "Flase" cuando una de las lineas no cumpla con el formato establecido. 
	
	if len(puntos) != 0:
			puntos = []#En caso se que el usuario vuelva a cargar informacion, se borrara la que estaba con anterioridad.
	
	if formato_elegido.get() == 1:
		
		dir_archivo = filedialog.askopenfilename(title="Abrir", initialdir = "desktop", filetypes = (("Ficheros de Texto", "*.txt"), ("Todos los Ficheros", "*.*")))	
	
		if dir_archivo == '':#Para evitar que salte un error si el usuario decide cerrar la interfas de seleccion de archivo.
			pass
	
		else:
			archivo = open(dir_archivo, "r")
			archivo = archivo.readlines()

			if not(archivo[0] == "AutoCAD-MIM por FeloCAD\n"):#Todos los archivos que genera el comando PM tienen esta linea al inicio.
				messagebox.showwarning("Nop", "Este archivo no procede del comando PM")

			else:
				for i in range(1, len(archivo)):#Desde uno para evitar la primera linea.
					archivo[i] = archivo[i].split()#sin argumentos separa por espacios.

				for j in range(1, len(archivo)):
					try:
						puntos.append(Punto(j, float(archivo[j][0]), float(archivo[j][1])))

					except:#No esefisique el error, porque puede ocurrir otros a demas de valueError.
						pasaron = False
						messagebox.showwarning(":(", "Error en la linea No.{}".format(j))
						puntos = []
						break #detener el ciclo ya que no tine sentido seguir si no se pudo cargar uno de los puntos.

	elif formato_elegido.get() == 2:

		dir_archivo = filedialog.askopenfilename(title="Abrir", initialdir = "desktop", filetypes = (("Ficheros de Texto", "*.txt"), ("Todos los Ficheros", "*.*")))	
	
		if not(dir_archivo == ''):
			archivo = open(dir_archivo, "r")
			archivo = archivo.readlines()

			for i in range(len(archivo)):
				EyN = extraerXY(archivo[i])
				j = i + 1 
				if  EyN == False:
					messagebox.showwarning(":/", "Error en la linea No.{}".format(j))
					pasaron = EyN
					puntos = []
					break
									
				else:	
					puntos.append(Punto(j, EyN[0], EyN[1]))
						
	else:
		messagebox.showinfo("Hey!", "Por favor elige el formato de entrada")
		pasaron = False#Esto evitara que entre al siguiente condicional en caso de que no se elijan ninguna de las opciones.


	if pasaron == True and puntos != []:

		while puntos[0].este == puntos[len(puntos)-1].este and puntos[0].norte == puntos[len(puntos)-1].norte:#uso un while porque esto puede pasar mas de una vez.
			puntos.pop()#en caso de que se repita el ultimo par de coordenadas borrar el ultimo punto.
			print("Hey!!")
				
		for p in puntos:
			print("E-{},E = {},N = {}".format(p.nombre, round(p.este, 2), round(p.norte, 2)))

		messagebox.showinfo(":)", "La información ha sido cargada con exisito")

def dar_formato_tabla(hoja):
	global puntos

	formato_millares = '#,##0.00'
	fuente = Font(name = "Arial", size = 11)
	linea_borde = Side(border_style = "thin")
	alineacion = Alignment(horizontal = "center")
	todos_los_bordes = Border(top = linea_borde, left = linea_borde, right = linea_borde, bottom = linea_borde)

	#Dar Formato a los encabezados:

	for fila in hoja.iter_rows(min_row = 3, max_row = 4, min_col = 2, max_col = 6):
		for celda in fila:
			celda.font = Font(name = "Arial", size = 12, bold = True)
			celda.border = todos_los_bordes
			celda.alignment = alineacion

	#Dar formato a las demas celdas de la tabla:

	for fila in hoja.iter_rows(min_row = 5, max_row = 4 + len(puntos), min_col = 2, max_col = 3):
		for celda in fila:
			celda.font, celda.alignment, celda.border = fuente, alineacion, todos_los_bordes

	for fila in hoja.iter_rows(min_row = 5, max_row = 4 + len(puntos), min_col = 4, max_col = 6):
		for celda in fila:
			celda.font, celda.alignment, celda.border, celda.number_format = fuente, alineacion, todos_los_bordes, formato_millares 			

	#Definir los anchos de las columnas:

	hoja.column_dimensions["B"].width = 10		
	hoja.column_dimensions["C"].width = 12
	hoja.column_dimensions["D"].width = 17
	hoja.column_dimensions["E"].width = 15
	hoja.column_dimensions["F"].width = 15
	hoja.column_dimensions["H"].width = 20
	hoja.column_dimensions["I"].width = 20
	hoja.column_dimensions["J"].width = 20

	#Para alimentar mi ego:

	hoja.merge_cells("B1:F1")
	hoja["B1"] = "Por Agrim Juan A. Núñez"
	hoja["B1"].font = Font(name = "times new roman", size = 12, bold = True, color = "000000FF")
	hoja["B1"].alignment = alineacion

	#Formato de las celdas del area y el perimetro:

	for casilla in ["H3", "H4"]:
		hoja[casilla].font = Font(size = 12, bold = True, color = "000000FF")
		hoja[casilla].alignment = alineacion

def calcular_area(puntos):
	area = 0

	for i in range(len(puntos)-1):
		area += puntos[i].este * puntos[i + 1].norte
		area -= puntos[i].norte * puntos[i + 1].este

	area += puntos[len(puntos) - 1].este * puntos[0].norte 	
	area -= puntos[len(puntos) - 1].norte * puntos[0].este

	area = 0.5*abs(area)

	if area > 628.86:
		return "Area = {} m^2 = {} tareas".format(round(area, 2), round(area/628.86, 2))

	else:
		return "Area = {} m^2".format(round(area, 2))

def inset_info():
	global puntos
	area = 0
	perimetro = 0

	if puntos == []:
		messagebox.showwarning("-_-", "hey, todavia no has cargado ningun archivo")
	
	else:
		direccion = filedialog.asksaveasfilename(defaultextension = ".xlsx", filetypes = (("Libro de Excel", "*.xlsx"), ("Todos los Ficheros", "*.*")))	

		if not(direccion == ''):
			libro = Workbook()
			hoja = libro.active
			hoja.title = "From Python"

#			hoja["A1"] = "Look at me"

			hoja.merge_cells("B3:F3")

			hoja["B3"] = "Coordenadas UTM Zona 19 Norte"
			hoja["B4"] = "Estación"
			hoja["C4"] = "Rumbo"
			hoja["D4"] = "Distancia(mts)"
			hoja["E4"] = "Este(x)"
			hoja["F4"] = "Norte(y)"

			#Definir las lineas:

			longitud = len(puntos)
			i = 0#Contador para iterar la lista de puntos.

			for fila in hoja.iter_rows(min_row = 5, max_row = 3 + longitud, min_col = 2, max_col = 6):
				ryd = rumbo_y_dist(puntos[i], puntos[i + 1])
				perimetro += ryd[1]
				fila[0].value = puntos[i].nombre
				fila[1].value = ryd[0]
				fila[2].value = round(ryd[1], 2)
				fila[3].value = round(puntos[i].este, 2)
				fila[4].value = round(puntos[i].norte, 2)
				i = i + 1
			
			#Definir la ultima linea:
				
			ryd_final = rumbo_y_dist(puntos[longitud - 1], puntos[0])#rumbo y distancia de la ultima linea a la primera.
			perimetro += ryd_final[1]	
			hoja.cell(row = 4 + longitud, column = 2, value = puntos[longitud - 1].nombre)	
			hoja.cell(row = 4 + longitud, column = 3, value = ryd_final[0])	
			hoja.cell(row = 4 + longitud, column = 4, value = round(ryd_final[1], 2))
			hoja.cell(row = 4 + longitud, column = 5, value = round(puntos[longitud - 1].este, 2))
			hoja.cell(row = 4 + longitud, column = 6, value = round(puntos[longitud - 1].norte, 2))

			#Introducir Area y Perimetro:
			hoja.merge_cells("H3:J3")
			hoja["H3"] = calcular_area(puntos)			
			hoja.merge_cells("H4:J4")
			hoja["H4"] = "Perimetro = {} metros".format(round(perimetro, 2))

			dar_formato_tabla(hoja)
			libro.save(direccion)
			messagebox.showinfo(":)", "La tabla ha sido guardada exitosamente en: {}".format(direccion))


#-----------------------------------------------------------------------------------Widgets---------------------------------------------------------------------------------------------

Etiketa = tk.Label(ventana, text = "Selecione Formato de Origen: ").pack(side = "top")

eleccion_pm = tk.Radiobutton(ventana, text = " Comando 'PM' ", variable = formato_elegido, value = 1).pack(side = "top")
eleccion_list = tk.Radiobutton(ventana, text = " Comando 'LIST' ", variable = formato_elegido, value = 2).pack(side = "top")

boton_cargar = tk.Button(ventana, text = "Cargar txt", command = cargar)
boton_cargar.place(x = 10, y = 80)

boton_instrucciones = tk.Button(ventana, text = "Instrucciones", command = mostrar_intrucciones)
boton_instrucciones.place(x = 79, y = 80)

boton_exportar = tk.Button(ventana, text = "Exportar Tabla", command = inset_info)
boton_exportar.place(x = 165, y = 80)


ventana.mainloop()
























